import os
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import logout, login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse 
from django.db.models import Count
from django.utils import timezone

# Formularios y Modelos
from .forms import RegistroClienteForm, EditarPerfilForm, UsuarioCreationForm, UsuarioUpdateForm
from .models import User
from agendamiento.models import Cita, Servicio 

# Generación de Reportes
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
# Barra busqueda avanzada
from django.db.models import Q

# ===================================================================
# PERMISOS Y UTILIDADES
# ===================================================================

def es_admin(user):
    """Verifica si el usuario tiene rol administrativo"""
    return user.is_authenticated and (getattr(user, 'rol', None) in ['ADMIN', 'JEFE'] or user.is_superuser)

# ===================================================================
# VISTA DINÁMICA: HOME / DASHBOARD
# ===================================================================

def home(request):
    if es_admin(request.user):
        total_usuarios = User.objects.count()
        total_citas = Cita.objects.count()
        hoy = timezone.now().date()
        citas_hoy = Cita.objects.filter(fecha_hora__date=hoy).count()
        
        citas_gestion = Cita.objects.all().order_by('-fecha_hora')
        ultimos_usuarios = User.objects.all().order_by('-id')[:5]

        context = {
            'total_usuarios': total_usuarios,
            'total_citas': total_citas,
            'citas_hoy': citas_hoy,
            'citas_gestion': citas_gestion,
            'ultimos_usuarios': ultimos_usuarios,
            'es_admin_dashboard': True  
        }
        return render(request, 'home.html', context)
    
    return render(request, 'home.html')

# ===================================================================
# GESTIÓN DE SESIÓN Y PERFIL
# ===================================================================

def custom_logout(request):
    logout(request)
    messages.info(request, "Sesión cerrada correctamente.")
    return redirect('home')

def registro_cliente(request):
    if request.method == 'POST':
        form = RegistroClienteForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, f"¡Bienvenido {user.username}! Registro exitoso.")
            return redirect('home')
    else:
        form = RegistroClienteForm()
    return render(request, 'registration/registro.html', {'form': form})

@login_required
def mi_perfil(request):
    if request.method == 'POST':
        form = EditarPerfilForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Perfil actualizado correctamente.")
            return redirect('mi_perfil')
    else:
        form = EditarPerfilForm(instance=request.user)
    return render(request, 'registration/perfil.html', {'form': form})

# ===================================================================
# ADMINISTRACIÓN DE USUARIOS (CRUD)
# ===================================================================

@login_required
@user_passes_test(es_admin, login_url='home')
def lista_usuarios(request):
    busqueda = request.GET.get('search')
    usuarios_list = User.objects.all().order_by('id') 
    if busqueda:
        usuarios_list = usuarios_list.filter(
            Q(username__icontains=busqueda) | 
            Q(first_name__icontains=busqueda) | 
            Q(last_name__icontains=busqueda) |
            Q(email__icontains=busqueda) |
            Q(identificacion__icontains=busqueda) |
            Q(telefono__icontains=busqueda) |
            Q(rol__icontains=busqueda)
        ).distinct()
    paginator = Paginator(usuarios_list, 10) 
    page_number = request.GET.get('page')
    usuarios = paginator.get_page(page_number)
    return render(request, 'usuarios/lista_usuarios.html', {
        'usuarios': usuarios, 
        'busqueda': busqueda
    })

@login_required
@user_passes_test(es_admin, login_url='home')
def crear_usuario(request):
    if request.method == 'POST':
        form = UsuarioCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Nuevo usuario registrado en el sistema.')
            return redirect('lista_usuarios')
    else:
        form = UsuarioCreationForm()
    return render(request, 'usuarios/form_usuario.html', {'form': form, 'titulo': 'Crear Usuario'})

@login_required
@user_passes_test(es_admin, login_url='home')
def editar_usuario(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = UsuarioUpdateForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, f'Datos de {usuario.username} actualizados.')
            return redirect('lista_usuarios')
    else:
        form = UsuarioUpdateForm(instance=usuario)
    return render(request, 'usuarios/form_usuario.html', {'form': form, 'titulo': 'Editar Usuario'})

@login_required
@user_passes_test(es_admin, login_url='home')
def eliminar_usuario(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        usuario.delete()
        messages.success(request, 'Usuario eliminado definitivamente.')
        return redirect('lista_usuarios')
    return render(request, 'usuarios/confirmar_eliminacion.html', {'usuario': usuario})

# ===================================================================
# EXPORTACIÓN DE REPORTES
# ===================================================================

@login_required
@user_passes_test(es_admin, login_url='home')
def generar_pdf_usuarios(request):
    scope = request.GET.get('scope', 'todo')
    
    if scope == 'pagina':
        usuarios_list = User.objects.all().order_by('id')
        paginator = Paginator(usuarios_list, 10)
        page_number = request.GET.get('page', 1)
        usuarios_reporte = paginator.get_page(page_number)
        subtitulo = f"Vista de Página {page_number}"
    else:
        usuarios_reporte = User.objects.all().order_by('id')
        subtitulo = "Base de Datos Completa"

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="EasyTime_{scope}.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    # Encabezado y Logo
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'favicon.ico')
    if os.path.exists(logo_path):
        p.drawImage(logo_path, 50, height - 75, width=40, height=40, mask='auto')
    
    p.setFont("Helvetica-Bold", 16)
    p.setFillColor(colors.HexColor("#0d6efd"))
    p.drawString(100, height - 55, "EASYTIME - Reporte de Usuarios")
    p.setFont("Helvetica", 9)
    p.setFillColor(colors.grey)
    p.drawString(100, height - 70, f"{subtitulo} | Generado por: {request.user.username}")

    # Tabla de datos
    headers = [['ID', 'Identificación', 'Nombre', 'Rol', 'Estado']]
    data = headers
    for user in usuarios_reporte:
        doc = getattr(user, 'identificacion', 'N/A') or 'N/A'
        nombre = f"{user.first_name} {user.last_name}" if user.first_name else user.username
        rol = getattr(user, 'rol', 'CLIENTE')
        estado = "Activo" if user.is_active else "Inactivo"
        data.append([user.id, doc, nombre, rol, estado])

    tabla = Table(data, colWidths=[40, 100, 180, 80, 80])
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ])
    
    # Filas alternas (Cebreado) corregido
    for i in range(1, len(data)):
        if i % 2 == 0:
            style.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)

    tabla.setStyle(style)
    w_t, h_t = tabla.wrap(width - 100, height)
    tabla.drawOn(p, 50, height - 120 - h_t)

    p.showPage()
    p.save()
    return response

@login_required
@user_passes_test(es_admin, login_url='home')
def generar_excel_usuarios(request):
    scope = request.GET.get('scope', 'todo')
    
    if scope == 'pagina':
        usuarios_list = User.objects.all().order_by('id')
        paginator = Paginator(usuarios_list, 10)
        page_number = request.GET.get('page', 1)
        usuarios_reporte = paginator.get_page(page_number)
        nombre_f = f"EasyTime_Pagina_{page_number}.xlsx"
    else:
        usuarios_reporte = User.objects.all().order_by('id')
        nombre_f = "EasyTime_Usuarios_Completo.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    ws.append(['ID', 'Identificación', 'Nombre', 'Apellido', 'Email', 'Rol', 'Estado'])

    # Estilo Encabezados
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")

    for u in usuarios_reporte:
        doc = getattr(u, 'identificacion', 'N/A') or 'N/A'
        ws.append([u.id, doc, u.first_name, u.last_name, u.email, getattr(u, 'rol', 'CLIENTE'), "Activo" if u.is_active else "Inactivo"])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_f}"'
    wb.save(response)
    return response
